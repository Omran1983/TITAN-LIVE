import os, time, hmac, hashlib, argparse
import urllib.parse as up
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import load_workbook
from scripts.env_loader import load_env

HEADERS = [
    "DateUTC","DateLocal","Env","Symbol","Side","OrderType","OrderId","ClientOrderId",
    "OrigQty","FilledQty","AvgFillPrice","Fee","FeeAsset","Status","Price",
    "StopPrice","TimeInForce","IsMaker"
]

def _env(k, d=""):
    v = os.getenv(k)
    return v if v not in (None, "") else d

def _sign(params: dict, secret: str) -> str:
    qs = up.urlencode(params, doseq=True)
    sig = hmac.new(secret.encode(), qs.encode(), hashlib.sha256).hexdigest()
    return f"{qs}&signature={sig}"

def _safe_save(wb, final_path: Path, attempts: int = 10, delay: float = 0.6) -> str:
    for i in range(attempts):
        try:
            wb.save(final_path)
            return str(final_path)
        except PermissionError:
            time.sleep(delay)
    # Fallback: save to a pending file so we don't lose data
    pending = final_path.with_suffix(".pending.xlsx")
    wb.save(pending)
    print(f"[excel] File locked. Saved to: {pending}. Close Excel and run finalize to swap.")
    return str(pending)

def fetch_mytrades_for_order(symbol: str, order_id: int):
    base = _env("BINANCE_BASE_URL", "https://api.binance.com")
    key  = _env("BINANCE_API_KEY")
    sec  = _env("BINANCE_API_SECRET")
    if not key or not sec:
        raise RuntimeError("Missing BINANCE_API_KEY / BINANCE_API_SECRET")

    ts = int(time.time() * 1000)
    params = {"symbol": symbol.upper(), "timestamp": ts, "recvWindow": 10000, "limit": 500}
    url = f"{base}/api/v3/myTrades?{_sign(params, sec)}"
    r = requests.get(url, headers={"X-MBX-APIKEY": key}, timeout=20)
    r.raise_for_status()
    trades = [t for t in r.json() if int(t.get("orderId", -1)) == int(order_id)]
    return trades

def enrich_excel(symbol: str, order_id: int, xlsx: str = "logs/trades.xlsx"):
    trades = fetch_mytrades_for_order(symbol, order_id)
    if not trades:
        print("No trades found for that orderId in myTrades (try again in ~30s).")
        return

    # Weighted avg price, qty, and fees
    tot_qty = 0.0
    tot_cost = 0.0
    fee_map = defaultdict(float)
    makers = set()

    for t in trades:
        q = float(t.get("qty", 0) or 0)
        p = float(t.get("price", 0) or 0)
        tot_qty += q
        tot_cost += q * p
        makers.add(bool(t.get("isMaker")))
        if "commission" in t and "commissionAsset" in t:
            fee_map[t["commissionAsset"]] += float(t.get("commission") or 0)

    avg_price = round(tot_cost / tot_qty, 8) if tot_qty > 0 else ""
    is_maker = "MAKER" if makers == {True} else ("TAKER" if makers == {False} else "MIXED")

    if len(fee_map) == 1:
        (asset, amt), = fee_map.items()
        fee_total = round(amt, 8)
        fee_asset = asset
    else:
        fee_total = "+".join(f"{round(v, 8)}{k}" for k, v in fee_map.items()) if fee_map else ""
        fee_asset = "MIXED" if len(fee_map) > 1 else ""

    xlsx_path = Path(xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"{xlsx} not found. Run a trade first so the logger creates it.")

    wb = load_workbook(xlsx_path)
    if "Trades" not in wb.sheetnames:
        raise ValueError("Trades sheet not found in workbook.")
    ws = wb["Trades"]

    order_idx = HEADERS.index("OrderId") + 1
    avg_idx   = HEADERS.index("AvgFillPrice") + 1
    fee_idx   = HEADERS.index("Fee") + 1
    feea_idx  = HEADERS.index("FeeAsset") + 1
    filled_idx= HEADERS.index("FilledQty") + 1
    ismk_idx  = HEADERS.index("IsMaker") + 1

    updated = False
    for row in ws.iter_rows(min_row=2):
        if str(row[order_idx - 1].value) == str(order_id):
            ws.cell(row=row[0].row, column=avg_idx, value=avg_price)
            ws.cell(row=row[0].row, column=filled_idx, value=round(tot_qty, 8))
            ws.cell(row=row[0].row, column=fee_idx, value=fee_total)
            ws.cell(row=row[0].row, column=feea_idx, value=fee_asset)
            ws.cell(row=row[0].row, column=ismk_idx, value=is_maker)
            updated = True
            break

    if not updated:
        raise ValueError("OrderId not found in Trades sheet. Did the order get logged?")

    saved = _safe_save(wb, xlsx_path)
    print(f"Enriched: {saved}")

def main():
    load_env()
    ap = argparse.ArgumentParser(description="Enrich a logged order with precise fills/fees via myTrades.")
    ap.add_argument("--symbol", required=True)
    ap.add_argument("--orderId", required=True, type=int)
    args = ap.parse_args()
    enrich_excel(args.symbol, args.orderId)

if __name__ == "__main__":
    main()
