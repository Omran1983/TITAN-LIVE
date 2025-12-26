from __future__ import annotations
import math
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import load_workbook, Workbook

TRADES_SHEET = "Trades"
CLOSED_SHEET = "ClosedTrades"
OPEN_SHEET   = "OpenPositions"

def _to_float(x, default=0.0):
    try:
        if x in ("", None): return default
        return float(x)
    except Exception:
        return default

def _safe_save(wb, final_path: Path, attempts: int = 10, delay: float = 0.5) -> str:
    import time
    for _ in range(attempts):
        try:
            wb.save(final_path)
            return str(final_path)
        except PermissionError:
            time.sleep(delay)
    pending = final_path.with_suffix(".pending.xlsx")
    wb.save(pending)
    print(f"[excel] Locked. Saved to: {pending}. Close Excel and run finalize if needed.")
    return str(pending)

def load_trades(xlsx="logs/trades.xlsx") -> List[Dict[str, Any]]:
    p = Path(xlsx)
    if not p.exists():
        raise FileNotFoundError(f"{xlsx} not found.")
    wb = load_workbook(p, data_only=True)
    if TRADES_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet {TRADES_SHEET} not found in {xlsx}")
    ws = wb[TRADES_SHEET]
    # headers
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in r):  # skip empty
            continue
        rows.append({
            "DateUTC": r[idx.get("DateUTC", -1)] if idx.get("DateUTC", -1) != -1 else "",
            "DateLocal": r[idx.get("DateLocal", -1)] if idx.get("DateLocal", -1) != -1 else "",
            "Env": r[idx.get("Env", -1)] if idx.get("Env", -1) != -1 else "",
            "Symbol": r[idx.get("Symbol", -1)] if idx.get("Symbol", -1) != -1 else "",
            "Side": r[idx.get("Side", -1)] if idx.get("Side", -1) != -1 else "",
            "OrderType": r[idx.get("OrderType", -1)] if idx.get("OrderType", -1) != -1 else "",
            "OrderId": r[idx.get("OrderId", -1)] if idx.get("OrderId", -1) != -1 else "",
            "ClientOrderId": r[idx.get("ClientOrderId", -1)] if idx.get("ClientOrderId", -1) != -1 else "",
            "OrigQty": _to_float(r[idx.get("OrigQty", -1)]) if idx.get("OrigQty", -1) != -1 else 0.0,
            "FilledQty": _to_float(r[idx.get("FilledQty", -1)]) if idx.get("FilledQty", -1) != -1 else 0.0,
            "AvgFillPrice": _to_float(r[idx.get("AvgFillPrice", -1)]) if idx.get("AvgFillPrice", -1) != -1 else 0.0,
            "Status": r[idx.get("Status", -1)] if idx.get("Status", -1) != -1 else "",
            "TimeInForce": r[idx.get("TimeInForce", -1)] if idx.get("TimeInForce", -1) != -1 else "",
        })
    # sort by DateUTC if present, else keep order
    def _dt(x):
        try:
            return datetime.fromisoformat(str(x).replace("Z",""))
        except Exception:
            return datetime.min
    rows.sort(key=lambda x: _dt(x["DateUTC"]))
    return rows, wb, p

def fifo_pnl(rows: List[Dict[str, Any]]):
    """
    Compute realized P&L per SELL using FIFO for USDT-quoted pairs.
    Returns (closed_trades, open_positions)
    """
    buy_queues: Dict[str, List[Dict[str, float]]] = {}  # symbol -> list of lots {qty, price}
    closed = []  # each: {Symbol, Qty, BuyAvgPrice, SellAvgPrice, RealizedPnL_USDT, BuyOrderIds, SellOrderId, CloseDate}
    for row in rows:
        sym = str(row["Symbol"] or "").upper()
        if not sym.endswith("USDT"):
            continue  # keep it simple: quote=P&L in USDT only
        side = str(row["Side"] or "").upper()
        qty  = _to_float(row["FilledQty"], 0.0)
        px   = _to_float(row["AvgFillPrice"], 0.0)
        if qty <= 0 or px <= 0:  # skip incomplete rows
            continue
        q = buy_queues.setdefault(sym, [])
        if side == "BUY":
            q.append({"qty": qty, "price": px, "orderId": row.get("OrderId")})
        elif side == "SELL":
            remain = qty
            buy_ids = []
            realized = 0.0
            matched_qty = 0.0
            while remain > 1e-15 and q:
                lot = q[0]
                use = min(lot["qty"], remain)
                realized += use * (px - lot["price"])
                matched_qty += use
                buy_ids.append(str(lot.get("orderId") or ""))
                lot["qty"] -= use
                remain -= use
                if lot["qty"] <= 1e-15:
                    q.pop(0)
            if matched_qty > 0:
                closed.append({
                    "CloseDate": row.get("DateLocal") or row.get("DateUTC"),
                    "Symbol": sym,
                    "Qty": round(matched_qty, 8),
                    "BuyAvgPrice": round((px - realized/matched_qty) , 8),  # inferred from pnl
                    "SellAvgPrice": round(px, 8),
                    "RealizedPnL_USDT": round(realized, 8),
                    "BuyOrderIds": ",".join([b for b in buy_ids if b]),
                    "SellOrderId": str(row.get("OrderId") or ""),
                })
            # if remain > 0 and no buys to match, we ignore (short sell not handled)
    # open positions
    opens = []
    for sym, lots in buy_queues.items():
        total_qty = sum(l["qty"] for l in lots)
        if total_qty <= 1e-15:
            continue
        cost = sum(l["qty"] * l["price"] for l in lots)
        opens.append({
            "Symbol": sym,
            "Qty": round(total_qty, 8),
            "AvgCost": round(cost/total_qty, 8)
        })
    return closed, opens

def write_outputs(wb, book_path: Path, closed: List[Dict[str, Any]], opens: List[Dict[str, Any]]):
    # nuke & recreate sheets
    for sh in (CLOSED_SHEET, OPEN_SHEET):
        if sh in wb.sheetnames:
            ws = wb[sh]
            wb.remove(ws)
    ws_c = wb.create_sheet(CLOSED_SHEET)
    ws_o = wb.create_sheet(OPEN_SHEET)

    # ClosedTrades
    c_headers = ["CloseDate","Symbol","Qty","BuyAvgPrice","SellAvgPrice","RealizedPnL_USDT","BuyOrderIds","SellOrderId"]
    ws_c.append(c_headers)
    for r in closed:
        ws_c.append([r.get(h,"") for h in c_headers])
    ws_c.freeze_panes = "A2"

    # OpenPositions
    o_headers = ["Symbol","Qty","AvgCost"]
    ws_o.append(o_headers)
    for r in opens:
        ws_o.append([r.get(h,"") for h in o_headers])
    ws_o.freeze_panes = "A2"

    saved = _safe_save(wb, book_path)
    print(f"[pnl] Wrote {CLOSED_SHEET} ({len(closed)} rows) and {OPEN_SHEET} ({len(opens)} rows) → {saved}")

def main():
    rows, wb, p = load_trades()
    closed, opens = fifo_pnl(rows)
    write_outputs(wb, p, closed, opens)

if __name__ == "__main__":
    main()
