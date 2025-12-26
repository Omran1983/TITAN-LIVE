from __future__ import annotations

import os
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict
from openpyxl import Workbook, load_workbook

HEADERS = [
    "DateUTC","DateLocal","Env","Symbol","Side","OrderType","OrderId","ClientOrderId",
    "OrigQty","FilledQty","AvgFillPrice","Fee","FeeAsset","Status","Price",
    "StopPrice","TimeInForce","IsMaker"
]

def _ensure_workbook(xlsx_path: Path) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if not xlsx_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    else:
        wb = load_workbook(xlsx_path)
        if "Trades" not in wb.sheetnames:
            ws = wb.create_sheet("Trades")
            ws.append(HEADERS)
            ws.freeze_panes = "A2"
            wb.save(xlsx_path)

def _order_id_exists(xlsx_path: Path, order_id: str) -> bool:
    if not xlsx_path.exists():
        return False
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Trades"]
    idx = HEADERS.index("OrderId")
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = row[idx]
        if cell is not None and str(cell) == str(order_id):
            return True
    return False

def log_order_fill(resp: Dict[str, Any], env: str = None, xlsx: str = "logs/trades.xlsx") -> str:
    """
    Append an order response to Excel.
    - Idempotent by OrderId.
    - If 'fills' missing, computes AvgFillPrice from cummulativeQuoteQty / executedQty when available.
    """
    xlsx_path = Path(xlsx)
    _ensure_workbook(xlsx_path)

    order_id = resp.get("orderId")
    if order_id and _order_id_exists(xlsx_path, order_id):
        return str(xlsx_path)

    now_utc = datetime.now(timezone.utc).isoformat()
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    symbol = resp.get("symbol") or ""
    side = resp.get("side") or ""
    otype = resp.get("type") or ""
    client_oid = resp.get("clientOrderId") or ""
    status = resp.get("status") or "NEW"

    def _to_float(v, default=0.0):
        try:
            if v in (None, ""):
                return default
            return float(v)
        except Exception:
            return default

    orig_qty = _to_float(resp.get("origQty"))
    filled_qty = _to_float(resp.get("executedQty"))
    price = resp.get("price")
    price = _to_float(price, default="") if isinstance(price, (int, float, str)) else ""

    stop_price = str(resp.get("stopPrice") or "")
    tif = resp.get("timeInForce") or ""

    # Compute avg fill + fees if fills present; else fallback to cummulativeQuoteQty/executedQty
    avg_price = ""
    fee_total = ""
    fee_asset = ""
    fills = resp.get("fills") or []

    if fills:
        tot_cost = 0.0
        tot_qty = 0.0
        fee_sum = 0.0
        fee_ccy = ""
        for f in fills:
            q = _to_float(f.get("qty") or f.get("quantity"))
            p = _to_float(f.get("price"))
            tot_cost += q * p
            tot_qty += q
            if "commission" in f:
                fee_sum += _to_float(f.get("commission"))
                fee_ccy = f.get("commissionAsset") or fee_ccy
        if tot_qty > 0:
            avg_price = round(tot_cost / tot_qty, 8)
            filled_qty = tot_qty or filled_qty
        if fee_sum > 0:
            fee_total = round(fee_sum, 8)
            fee_asset = fee_ccy
    else:
        cum_quote = _to_float(resp.get("cummulativeQuoteQty"))
        if filled_qty and filled_qty > 0 and cum_quote > 0:
            avg_price = round(cum_quote / filled_qty, 8)

    wb = load_workbook(xlsx_path)
    ws = wb["Trades"]
    row = [
        now_utc,
        now_local,
        env or os.getenv("ENV_FILE") or "",
        symbol,
        side,
        otype,
        str(order_id or ""),
        client_oid,
        orig_qty,
        filled_qty,
        avg_price,
        fee_total,
        fee_asset,
        status,
        price,
        stop_price,
        tif,
        ""  # IsMaker (not on order response; available via myTrades)
    ]
    ws.append(row)
    wb.save(xlsx_path)
    return str(xlsx_path)
