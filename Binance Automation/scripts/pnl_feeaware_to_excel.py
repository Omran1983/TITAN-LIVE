from __future__ import annotations
import math, time
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, List, Any

import requests
from openpyxl import load_workbook

TRADES_SHEET = "Trades"
CLOSED_SHEET = "ClosedTrades"
OPEN_SHEET   = "OpenPositions"
DAILY_SHEET  = "DailyPnL"

def _to_f(v, d=0.0):
    try:
        if v in ("", None): return d
        return float(v)
    except: return d

def _parse_dt(x):
    # accepts ISO DateUTC or local string
    s = str(x or "")
    try:
        return datetime.fromisoformat(s.replace("Z",""))
    except:
        try:
            return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        except:
            return None

# --- price converters (at trade time) ---
BASE = "https://api.binance.com"
def _kl_close(symbol: str, ts_ms: int):
    # 1m candle around the timestamp
    try:
        params = {"symbol": symbol, "interval": "1m", "startTime": ts_ms-60_000, "endTime": ts_ms+60_000, "limit": 1}
        r = requests.get(f"{BASE}/api/v3/klines", params=params, timeout=10)
        r.raise_for_status()
        arr = r.json()
        if arr:
            return float(arr[0][4])  # close
    except: pass
    # fallback: spot now
    try:
        r2 = requests.get(f"{BASE}/api/v3/ticker/price", params={"symbol": symbol}, timeout=10)
        r2.raise_for_status()
        return float(r2.json()["price"])
    except: return None

_price_cache: Dict[tuple, float] = {}

def convert_fee_to_usdt(asset: str, amt: float, ts_ms: int) -> float:
    if not asset or amt == 0: return 0.0
    if asset.upper() in ("USDT", "FDUSD", "BUSD"):
        return amt
    sym = f"{asset.upper()}USDT"
    key = (sym, ts_ms//60_000)  # minute-bucket cache
    if key not in _price_cache:
        px = _kl_close(sym, ts_ms)
        _price_cache[key] = px or 0.0
    px = _price_cache[key]
    return round(amt * px, 8) if px else 0.0

def _safe_save(wb, path: Path):
    import time
    for _ in range(10):
        try:
            wb.save(path); return str(path)
        except PermissionError:
            time.sleep(0.5)
    pend = path.with_suffix(".pending.xlsx")
    wb.save(pend)
    return str(pend)

def load_ledger(path="logs/trades.xlsx"):
    p = Path(path)
    if not p.exists(): raise FileNotFoundError(path)
    wb = load_workbook(p, data_only=True)
    if TRADES_SHEET not in wb.sheetnames: raise ValueError("Trades sheet missing")
    ws = wb[TRADES_SHEET]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h:i for i,h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v in ("", None) for v in r): continue
        rows.append({
            "DateUTC": r[H.get("DateUTC", -1)],
            "DateLocal": r[H.get("DateLocal", -1)],
            "Env": r[H.get("Env", -1)],
            "Symbol": r[H.get("Symbol", -1)],
            "Side": r[H.get("Side", -1)],
            "OrderType": r[H.get("OrderType", -1)],
            "OrderId": r[H.get("OrderId", -1)],
            "ClientOrderId": r[H.get("ClientOrderId", -1)],
            "OrigQty": _to_f(r[H.get("OrigQty",-1)]),
            "FilledQty": _to_f(r[H.get("FilledQty",-1)]),
            "AvgFillPrice": _to_f(r[H.get("AvgFillPrice",-1)]),
            "Fee": r[H.get("Fee",-1)],
            "FeeAsset": r[H.get("FeeAsset",-1)],
            "Status": r[H.get("Status",-1)],
        })
    # order by timestamp for FIFO
    rows.sort(key=lambda x: (_parse_dt(x["DateUTC"]) or datetime.min))
    return rows, wb, p

def compute_feeaware_fifo(rows: List[Dict[str, Any]]):
    """
    Realized P&L in USDT using FIFO, subtracting fees on both legs (converted to USDT at trade time).
    """
    open_lots: Dict[str, List[Dict[str,float]]] = {}
    closed = []
    for row in rows:
        sym = str(row["Symbol"] or "").upper()
        if not sym.endswith("USDT"):   # scope: USDT-quoted pairs
            continue
        dt = _parse_dt(row["DateUTC"]) or _parse_dt(row["DateLocal"]) or datetime.now()
        ts_ms = int(dt.replace(tzinfo=timezone.utc).timestamp()*1000)
        side = str(row["Side"] or "").upper()
        qty  = _to_f(row["FilledQty"])
        px   = _to_f(row["AvgFillPrice"])
        # fee from ledger row (may be "", or number), convert to USDT
        fee_amt = _to_f(row.get("Fee"))
        fee_ccy = (row.get("FeeAsset") or "").strip()
        fee_usdt = convert_fee_to_usdt(fee_ccy, fee_amt, ts_ms)

        if qty <= 0 or px <= 0: 
            continue

        q = open_lots.setdefault(sym, [])
        if side == "BUY":
            q.append({"qty": qty, "price": px, "fee_usdt": fee_usdt})
        else:  # SELL
            remain = qty
            realized = 0.0
            matched = 0.0
            cost_side_fees = 0.0
            while remain > 1e-15 and q:
                lot = q[0]
                use = min(lot["qty"], remain)
                realized += use * (px - lot["price"])
                matched += use
                # allocate buy-fee proportionally
                if lot["qty"] > 0:
                    alloc = lot["fee_usdt"] * (use / lot["qty"])
                    cost_side_fees += alloc
                lot["qty"] -= use
                remain -= use
                if lot["qty"] <= 1e-15:
                    q.pop(0)
            # subtract SELL fees and allocated BUY fees
            sell_fee_usdt = fee_usdt
            net_realized = realized - cost_side_fees - sell_fee_usdt
            if matched > 0:
                closed.append({
                    "CloseDate": row.get("DateLocal") or row.get("DateUTC"),
                    "Symbol": sym,
                    "Qty": round(matched, 8),
                    "BuyAvgPrice": round((px - realized/matched), 8),
                    "SellAvgPrice": round(px, 8),
                    "GrossPnL_USDT": round(realized, 8),
                    "Fees_USDT": round(cost_side_fees + sell_fee_usdt, 8),
                    "RealizedPnL_USDT": round(net_realized, 8),
                })

    # open positions
    opens = []
    for sym, lots in open_lots.items():
        tot = sum(l["qty"] for l in lots)
        if tot <= 1e-15: continue
        cost = sum(l["qty"]*l["price"] for l in lots)
        fee  = sum(l["fee_usdt"]       for l in lots)
        opens.append({"Symbol": sym, "Qty": round(tot,8), "AvgCost": round(cost/tot,8), "BuyFees_USDT": round(fee,8)})

    # daily rollup from closed
    daily: Dict[str, Dict[str,float]] = {}
    for r in closed:
        day = str(r["CloseDate"])[:10]
        d = daily.setdefault(day, {"RealizedPnL_USDT":0.0, "Fees_USDT":0.0, "GrossPnL_USDT":0.0, "Trades":0})
        d["RealizedPnL_USDT"] += r["RealizedPnL_USDT"]
        d["Fees_USDT"] += r["Fees_USDT"]
        d["GrossPnL_USDT"] += r["GrossPnL_USDT"]
        d["Trades"] += 1

    daily_rows = [{"Date":k, **v} for k,v in sorted(daily.items())]
    return closed, opens, daily_rows

def write_sheets(wb, book_path: Path, closed, opens, daily_rows):
    # drop old sheets if exist
    for sh in (CLOSED_SHEET, OPEN_SHEET, DAILY_SHEET):
        if sh in wb.sheetnames:
            wb.remove(wb[sh])
    ws_c = wb.create_sheet(CLOSED_SHEET)
    ws_o = wb.create_sheet(OPEN_SHEET)
    ws_d = wb.create_sheet(DAILY_SHEET)

    # ClosedTrades
    c_headers = ["CloseDate","Symbol","Qty","BuyAvgPrice","SellAvgPrice","GrossPnL_USDT","Fees_USDT","RealizedPnL_USDT"]
    ws_c.append(c_headers)
    for r in closed:
        ws_c.append([r.get(h,"") for h in c_headers])
    ws_c.freeze_panes = "A2"

    # OpenPositions
    o_headers = ["Symbol","Qty","AvgCost","BuyFees_USDT"]
    ws_o.append(o_headers)
    for r in opens:
        ws_o.append([r.get(h,"") for h in o_headers])
    ws_o.freeze_panes = "A2"

    # DailyPnL
    d_headers = ["Date","Trades","GrossPnL_USDT","Fees_USDT","RealizedPnL_USDT"]
    ws_d.append(d_headers)
    for r in daily_rows:
        ws_d.append([r.get(h,"") for h in d_headers])
    ws_d.freeze_panes = "A2"

    path = _safe_save(wb, book_path)
    print(f"[pnl] Wrote {len(closed)} closed, {len(opens)} open, {len(daily_rows)} days → {path}")

def main():
    rows, wb, p = load_ledger()
    closed, opens, daily = compute_feeaware_fifo(rows)
    write_sheets(wb, p, closed, opens, daily)

if __name__ == "__main__":
    main()
