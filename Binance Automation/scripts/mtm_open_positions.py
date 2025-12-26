from pathlib import Path
from openpyxl import load_workbook
import requests

BASE = "https://api.binance.com"

def spot(symbol):
    r = requests.get(f"{BASE}/api/v3/ticker/price", params={"symbol":symbol}, timeout=10)
    r.raise_for_status()
    return float(r.json()["price"])

def update_open_positions(path="logs/trades.xlsx"):
    p = Path(path); wb = load_workbook(p)
    if "OpenPositions" not in wb.sheetnames:
        print("OpenPositions sheet not found. Run pnl_feeaware_to_excel first.")
        return
    ws = wb["OpenPositions"]
    # header
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h:i+1 for i,h in enumerate(hdr)}
    # add new columns if missing
    for col in ("MarkPrice","UnrealizedPnL_USDT"):
        if col not in col_map:
            ws.cell(row=1, column=len(hdr)+1, value=col); hdr.append(col)
            col_map[col] = len(hdr)
    # update rows
    for row in ws.iter_rows(min_row=2):
        sym = str(row[col_map["Symbol"]-1].value or "")
        qty = float(row[col_map["Qty"]-1].value or 0)
        avg = float(row[col_map["AvgCost"]-1].value or 0)
        if not sym or qty<=0 or avg<=0: 
            continue
        try:
            mp = spot(sym)
            upnl = (mp - avg) * qty
            row[col_map["MarkPrice"]-1].value = mp
            row[col_map["UnrealizedPnL_USDT"]-1].value = round(upnl, 8)
        except Exception:
            pass
    wb.save(p)
    print(f"Updated MTM → {p}")

if __name__ == "__main__":
    update_open_positions()
