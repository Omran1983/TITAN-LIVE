from pathlib import Path
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook

HEADERS = [
    "DateUTC","DateLocal","Env","Symbol","Side","OrderType","OrderId","ClientOrderId",
    "OrigQty","FilledQty","AvgFillPrice","Fee","FeeAsset","Status","Price",
    "StopPrice","TimeInForce","IsMaker"
]

def ensure_book(p: Path):
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        wb.save(p)

def safe_save(wb, p: Path):
    import time
    for _ in range(10):
        try:
            wb.save(p); return str(p)
        except PermissionError:
            time.sleep(0.5)
    pend = p.with_suffix(".pending.xlsx")
    wb.save(pend); return str(pend)

def seed(symbol: str, qty: float, price_usdt: float, xlsx="logs/trades.xlsx"):
    p = Path(xlsx)
    ensure_book(p)
    wb = load_workbook(p)
    ws = wb["Trades"]

    now_utc = datetime.now(timezone.utc).isoformat()
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    oid = f"SEED-{symbol}-{int(datetime.now().timestamp())}"

    row = [
        now_utc, now_local, "SEED", symbol.upper(), "BUY", "SEED", oid, "",
        qty, qty, price_usdt, "", "", "FILLED", "", "", "", "SEED"
    ]
    ws.append(row)
    path = safe_save(wb, p)
    print(f"Seeded opening lot for {symbol}: qty={qty}, price={price_usdt} → {path}")

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Seed an opening lot (synthetic BUY) for deposited assets.")
    ap.add_argument("--symbol", required=True, help="e.g., BNBUSDT")
    ap.add_argument("--qty", required=True, type=float, help="Base asset qty (e.g., 0.25)")
    ap.add_argument("--price", required=True, type=float, help="Unit price in USDT (your cost basis)")
    args = ap.parse_args()
    seed(args.symbol, args.qty, args.price)
