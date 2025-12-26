import os, time, hmac, hashlib, argparse, math, json, re
import urllib.parse as up
from datetime import datetime, timedelta, timezone
from pathlib import Path
from collections import defaultdict

import requests
from openpyxl import Workbook, load_workbook
from scripts.env_loader import load_env

HEADERS = [
    "DateUTC","DateLocal","Env","Symbol","Side","OrderType","OrderId","ClientOrderId",
    "OrigQty","FilledQty","AvgFillPrice","Fee","FeeAsset","Status","Price",
    "StopPrice","TimeInForce","IsMaker"
]

TIME_OFFSET_MS = 0

def _env(k, d=""):
    v = os.getenv(k)
    return v if v not in (None, "") else d

def _base():
    return _env("BINANCE_BASE_URL", "https://api.binance.com")

def _key():
    return _env("BINANCE_API_KEY")

def _sec():
    return _env("BINANCE_API_SECRET")

def _server_time():
    r = requests.get(f"{_base()}/api/v3/time", timeout=10)
    r.raise_for_status()
    return int(r.json()["serverTime"])

def _sync_time():
    global TIME_OFFSET_MS
    try:
        st = _server_time()
        lt = int(time.time()*1000)
        TIME_OFFSET_MS = st - lt
        print(f"[time] offset ms: {TIME_OFFSET_MS}")
    except Exception as e:
        print(f"[time] WARN failed to sync time: {e}")

def _ts():
    return int(time.time()*1000 + TIME_OFFSET_MS)

def _sign(params: dict, secret: str) -> str:
    qs = up.urlencode(params, doseq=True)
    sig = hmac.new(secret.encode(), qs.encode(), hashlib.sha256).hexdigest()
    return f"{qs}&signature={sig}"

def _req(method: str, path: str, params: dict, signed: bool = False):
    if signed:
        params = dict(params or {})
        params.setdefault("timestamp", _ts())
        params.setdefault("recvWindow", 15000)
        url = f"{_base()}{path}?{_sign(params, _sec())}"
        headers = {"X-MBX-APIKEY": _key()}
    else:
        url = f"{_base()}{path}"
        headers = {}
    r = requests.request(method, url, headers=headers, timeout=30)
    if r.status_code >= 400:
        # Try to surface Binance error payload for fast triage
        try:
            payload = r.json()
        except Exception:
            payload = {"raw": r.text}
        raise requests.HTTPError(f"{r.status_code} {path} → {payload}", response=r)
    return r

def _ensure_workbook(xlsx_path: Path) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if not xlsx_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    else:
        wb = load_workbook(xlsx_path)
        if "Trades" not in wb.sheetnames:
            ws = wb.create_sheet("Trades")
            ws.append(HEADERS)
            ws.freeze_panes = "A2"
            wb.save(xlsx_path)

def _safe_save(wb, final_path: Path, attempts: int = 10, delay: float = 0.6) -> str:
    for _ in range(attempts):
        try:
            wb.save(final_path)
            return str(final_path)
        except PermissionError:
            time.sleep(delay)
    pending = final_path.with_suffix(".pending.xlsx")
    wb.save(pending)
    print(f"[excel] Locked. Saved to: {pending}. Close Excel then run finalize (scripts.finalize_excel_swap).")
    return str(pending)

def _exchange_info():
    r = _req("GET", "/api/v3/exchangeInfo", params={}, signed=False)
    data = r.json()
    symbols = {s["symbol"]: s for s in data.get("symbols", []) if s.get("status") == "TRADING"}
    return symbols

EXCLUDE_PAT = re.compile(r"(UPUSDT|DOWNUSDT|BULLUSDT|BEARUSDT)$")

def _account_balances():
    r = _req("GET", "/api/v3/account", params={}, signed=True)
    bals = r.json().get("balances", [])
    assets = []
    for b in bals:
        free = float(b.get("free", 0) or 0)
        locked = float(b.get("locked", 0) or 0)
        if (free + locked) > 0:
            assets.append(b.get("asset"))
    return assets

def _discover_symbols(quote: str, explicit_symbols: list[str], scan_exchange: bool) -> list[str]:
    info = _exchange_info()
    universe = set()

    # Explicit
    for s in explicit_symbols:
        if s in info and not EXCLUDE_PAT.search(s):
            universe.add(s)

    # From balances (ASSET+QUOTE)
    assets = _account_balances()
    for a in assets:
        sym = f"{a}{quote}".upper()
        if sym in info and not EXCLUDE_PAT.search(sym):
            universe.add(sym)

    # Broad scan (optional)
    if scan_exchange:
        for sym in info:
            if sym.endswith(quote.upper()) and not EXCLUDE_PAT.search(sym):
                universe.add(sym)

    return sorted(universe)

def _probe_symbol(symbol: str) -> bool:
    """Check quickly if myTrades is callable for this symbol. Returns True if OK."""
    try:
        params = {"symbol": symbol, "limit": 1}
        r = _req("GET", "/api/v3/myTrades", params=params, signed=True)
        _ = r.json()  # may be [] if no trades — that's fine
        return True
    except requests.HTTPError as e:
        msg = str(e)
        # Common causes:
        # -1121 Invalid symbol → skip
        # -1021 Timestamp outside recvWindow → resync time and retry once
        if "-1021" in msg or "Timestamp" in msg:
            print(f"[probe] {symbol}: timestamp issue, resyncing time and retrying once...")
            _sync_time()
            try:
                r = _req("GET", "/api/v3/myTrades", params={"symbol": symbol, "limit": 1}, signed=True)
                _ = r.json()
                return True
            except requests.HTTPError as e2:
                print(f"[probe] {symbol}: still failing → {e2}")
                return False
        elif "-1121" in msg or "Invalid symbol" in msg:
            print(f"[probe] {symbol}: invalid symbol, skipping.")
            return False
        else:
            print(f"[probe] {symbol}: unexpected error, skipping → {msg}")
            return False

def _mytrades_window(symbol: str, start_ms: int, end_ms: int) -> list[dict]:
    params = {
        "symbol": symbol,
        "startTime": start_ms,
        "endTime": end_ms,
        "limit": 1000
    }
    try:
        r = _req("GET", "/api/v3/myTrades", params=params, signed=True)
        return r.json()
    except requests.HTTPError as e:
        msg = str(e)
        if "-1021" in msg or "Timestamp" in msg:
            print(f"[myTrades] {symbol}: timestamp issue → resync & retry once.")
            _sync_time()
            r2 = _req("GET", "/api/v3/myTrades", params=params, signed=True)
            return r2.json()
        elif "-1121" in msg or "Invalid symbol" in msg:
            print(f"[myTrades] {symbol}: invalid symbol, skipping window.")
            return []
        else:
            # Some symbols simply return 400 when no trades with start/endTime; fall back to no window
            try:
                r3 = _req("GET", "/api/v3/myTrades", params={"symbol": symbol, "limit": 1000}, signed=True)
                return r3.json()
            except Exception:
                print(f"[myTrades] {symbol}: {msg}")
                return []

def _hydrate_order(symbol: str, order_id: int) -> dict | None:
    params = {"symbol": symbol, "orderId": int(order_id)}
    try:
        r = _req("GET", "/api/v3/order", params=params, signed=True)
        return r.json()
    except requests.HTTPError as e:
        print(f"[order] {symbol} {order_id}: {e}")
        return None

def _append_or_update_order(ws, env_name: str, symbol: str, agg: dict, details: dict | None):
    order_idx = HEADERS.index("OrderId") + 1
    found_row = None
    for row in ws.iter_rows(min_row=2):
        if str(row[order_idx-1].value) == str(agg["orderId"]):
            found_row = row[0].row
            break

    now_utc = datetime.now(timezone.utc).isoformat()
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    side = "BUY" if agg.get("isBuyer") else "SELL"
    avg_price = agg.get("avgPrice") or ""
    fee_total = agg.get("feeTotal") or ""
    fee_asset = agg.get("feeAsset") or ("MIXED" if isinstance(fee_total, str) and "+" in fee_total else "")
    filled_qty = agg.get("qty") or ""
    status = details.get("status") if details else "FILLED"
    otype = details.get("type") if details else ""
    tif = details.get("timeInForce") if details else ""
    price = details.get("price") if details else ""
    stop_price = details.get("stopPrice") if details else ""
    client_oid = details.get("clientOrderId") if details else ""
    orig_qty = details.get("origQty") if details else ""

    is_maker = agg.get("role")  # MAKER/TAKER/MIXED

    row_values = [
        now_utc, now_local, env_name, symbol, side, otype, str(agg["orderId"]), client_oid,
        orig_qty, filled_qty, avg_price, fee_total, fee_asset, status, price, stop_price, tif, is_maker
    ]

    if found_row:
        for ci, v in enumerate(row_values, start=1):
            ws.cell(row=found_row, column=ci, value=v)
    else:
        ws.append(row_values)

def _aggregate_order(trades: list[dict]) -> dict:
    qty = 0.0
    cost = 0.0
    fee_map = defaultdict(float)
    makers = set()
    is_buyer = bool(trades[0].get("isBuyer"))

    for t in trades:
        q = float(t.get("qty") or 0)
        p = float(t.get("price") or 0)
        qty += q
        cost += q * p
        makers.add(bool(t.get("isMaker")))
        if "commission" in t and "commissionAsset" in t:
            fee_map[t["commissionAsset"]] += float(t.get("commission") or 0)

    avg_price = round(cost/qty, 8) if qty > 0 else ""
    if len(fee_map) == 1:
        (asset, amt), = fee_map.items()
        fee_total = round(amt, 8)
        fee_asset = asset
    elif len(fee_map) > 1:
        fee_total = "+".join(f"{round(v,8)}{k}" for k, v in fee_map.items())
        fee_asset = "MIXED"
    else:
        fee_total, fee_asset = "", ""

    role = "MAKER" if makers == {True} else ("TAKER" if makers == {False} else "MIXED")

    return {
        "orderId": int(trades[0]["orderId"]),
        "avgPrice": avg_price,
        "qty": round(qty, 8) if qty else "",
        "feeTotal": fee_total,
        "feeAsset": fee_asset,
        "role": role,
        "isBuyer": is_buyer,
    }

def backfill(days: int, quote: str, symbols: list[str], scan_exchange: bool, hydrate_orders: bool, xlsx="logs/trades.xlsx", step_days=7):
    env_name = _env("ENV_FILE", ".env.mainnet")
    _ensure_workbook(Path(xlsx))
    _sync_time()  # once up-front

    info = _exchange_info()  # connectivity check
    candidates = _discover_symbols(quote, [s.upper() for s in symbols], scan_exchange)
    if not candidates:
        print("No symbols discovered. Pass --symbols or enable --scan-exchange.")
        return

    # Pre-filter by probe to avoid noisy 400s
    probed = [s for s in candidates if _probe_symbol(s)]
    if not probed:
        print("No probeable symbols. Check API permissions (Spot & Margin Trading) and IP whitelist.")
        return

    start_dt = datetime.now(timezone.utc) - timedelta(days=days)
    end_dt = datetime.now(timezone.utc)

    wb = load_workbook(xlsx)
    ws = wb["Trades"]

    for sym in probed:
        print(f"[backfill] {sym} ...")
        s = start_dt
        while s < end_dt:
            e = min(s + timedelta(days=step_days), end_dt)
            trades = _mytrades_window(sym, int(s.timestamp()*1000), int(e.timestamp()*1000))
            if trades:
                buckets = defaultdict(list)
                for t in trades:
                    buckets[int(t["orderId"])].append(t)
                for order_id, ts in buckets.items():
                    agg = _aggregate_order(ts)
                    details = _hydrate_order(sym, order_id) if hydrate_orders else None
                    _append_or_update_order(ws, env_name, sym, agg, details)
            time.sleep(0.12)  # be nice to rate limits
            s = e

        _safe_save(wb, Path(xlsx))

    print(f"[done] Backfill complete → {xlsx}")

def main():
    load_env()
    ap = argparse.ArgumentParser(description="Backfill all trades into logs/trades.xlsx, aggregated per orderId.")
    ap.add_argument("--days", type=int, default=365, help="How many days back to pull (default 365).")
    ap.add_argument("--quote", default="USDT", help="Quote currency to target (default USDT).")
    ap.add_argument("--symbols", default="", help="Comma-separated list like BTCUSDT,ETHUSDT,BNBUSDT.")
    ap.add_argument("--scan-exchange", action="store_true", help="Scan ALL TRADING pairs with the given quote (comprehensive, slower).")
    ap.add_argument("--no-hydrate-orders", action="store_true", help="Skip order detail hydration to reduce requests.")
    args = ap.parse_args()

    syms = [s.strip().upper() for s in args.symbols.split(",") if s.strip()]
    backfill(
        days=args.days,
        quote=args.quote.upper(),
        symbols=syms,
        scan_exchange=args.scan_exchange,
        hydrate_orders=(not args.no_hydrate_orders),
        xlsx="logs/trades.xlsx",
        step_days=7
    )

if __name__ == "__main__":
    main()
