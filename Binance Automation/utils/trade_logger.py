from __future__ import annotations
import os, threading
from datetime import datetime
from typing import Optional, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill

_HEADERS = [
    "Timestamp","Symbol","Side","OrderType","QtyBase","Price","QuoteAmt",
    "EntryPrice","ExitPrice","PnL_USDT","PnL_Pct","Cumulative_PnL",
    "Strategy","Sentiment","Notes","OrderId","OcoListId"
]

class TradeLogger:
    def __init__(self, xlsx_path: str = None):
        self.path = xlsx_path or os.path.join("logs","trades.xlsx")
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        self._lock = threading.Lock()
        if not os.path.exists(self.path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trades"
            for i,h in enumerate(_HEADERS,1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            wb.save(self.path)

    def log(self, row: Dict[str, Any]) -> None:
        with self._lock:
            wb = openpyxl.load_workbook(self.path)
            ws = wb["Trades"]
            out = [
                row.get("timestamp", datetime.now()),
                row.get("symbol"), row.get("side"), row.get("order_type"),
                row.get("qty_base"), row.get("price"), row.get("quote_amount"),
                row.get("entry_price"), row.get("exit_price"),
                row.get("pnl_usdt"), row.get("pnl_pct"),
                row.get("cumulative_pnl"),
                row.get("strategy"), row.get("sentiment"), row.get("notes"),
                row.get("orderId"), row.get("ocoListId")
            ]
            r = ws.max_row + 1
            for i,v in enumerate(out,1):
                ws.cell(row=r, column=i, value=v)
            # simple coloring on realized PnL rows
            pnl = (row.get("pnl_usdt") or 0) or 0
            if pnl != 0:
                fill = PatternFill(start_color=("00FF00" if pnl>0 else "FF0000"),
                                   end_color=("00FF00" if pnl>0 else "FF0000"),
                                   fill_type="solid")
                for col in range(1, len(_HEADERS)+1):
                    ws.cell(row=r, column=col).fill = fill
            wb.save(self.path)

# one shared instance
LOGGER = TradeLogger()
